"""給与明細印刷用 Excel（A4縦・1人1シート）を生成する。"""

import re
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

THIN = Side(style="thin")
BLOCK_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16)
NET_FONT = Font(bold=True, size=14)
NORMAL = Font(size=11)
AMOUNT_FMT = "#,##0"
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _safe_sheet_name(name: str, index: int) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "", (name or "").strip())
    cleaned = cleaned[:31]
    return cleaned or f"明細{index}"


def _unique_sheet_names(records):
    used = set()
    names = []
    for i, rec in enumerate(records, start=1):
        base = _safe_sheet_name(str(rec.get("氏名", "")), i)
        candidate = base
        suffix = 2
        while candidate in used:
            tail = f"_{suffix}"
            candidate = f"{base[: 31 - len(tail)]}{tail}"
            suffix += 1
        used.add(candidate)
        names.append(candidate)
    return names


def _deduction_total(record: dict) -> int:
    return int(record.get("D_社会保険料計", 0)) + int(record.get("F_控除計", 0))


def _apply_outer_border(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    """指定範囲の外周だけに黒罫線を付ける（内側の横線は不要）。"""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            left = THIN if c == start_col else None
            right = THIN if c == end_col else None
            top = THIN if r == start_row else None
            bottom = THIN if r == end_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _write_row(ws, row: int, label: str, amount: Optional[int], *, bold=False, net=False):
    label_font = Font(bold=bold, size=14 if net else 11)
    amount_font = NET_FONT if net else (Font(bold=bold, size=11))
    ws.cell(row=row, column=1, value=label).font = label_font
    ws.cell(row=row, column=1).alignment = LEFT
    ws.cell(row=row, column=1).border = BLOCK_BORDER
    if amount is not None:
        cell = ws.cell(row=row, column=2, value=amount)
        cell.number_format = AMOUNT_FMT
        cell.font = amount_font
        cell.alignment = RIGHT
    else:
        cell = ws.cell(row=row, column=2, value="")
    cell.border = BLOCK_BORDER


def _write_block_header(ws, row: int, title: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = HEADER_FONT
    cell.alignment = LEFT
    cell.border = BLOCK_BORDER
    ws.cell(row=row, column=2).border = BLOCK_BORDER


def _write_info_row(ws, row: int, label: str, value: str):
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = NORMAL
    label_cell.alignment = LEFT
    label_cell.border = BLOCK_BORDER
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = NORMAL
    value_cell.alignment = LEFT
    value_cell.border = BLOCK_BORDER


def _get_amount(record: dict, *keys, default=0):
    for key in keys:
        if key in record:
            return int(record[key])
    return default


def _pay_items(record: dict):
    """支給項目（新形式）。旧形式の記録にも対応。"""
    if "家族手当" in record:
        return [
            ("基本給", _get_amount(record, "基本給")),
            ("時間外手当", _get_amount(record, "時間外手当")),
            ("家族手当", _get_amount(record, "家族手当")),
            ("通勤手当", _get_amount(record, "通勤手当")),
            ("資格手当", _get_amount(record, "資格手当")),
            ("通信手当", _get_amount(record, "通信手当")),
        ]
    return [
        ("基本給", _get_amount(record, "基本給")),
        ("時間外手当", _get_amount(record, "時間外手当")),
        ("家族手当", _get_amount(record, "通勤手当")),
        ("通勤手当", _get_amount(record, "その他手当1")),
        ("資格手当", _get_amount(record, "その他手当2")),
        ("通信手当", _get_amount(record, "その他手当3")),
    ]


def _build_sheet(ws, record: dict):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    title = ws.cell(row=row, column=1, value="給与明細書")
    title.font = TITLE_FONT
    title.alignment = CENTER

    row = 3
    _write_info_row(ws, row, "月分", str(record.get("月分", "")))
    row += 1
    _write_info_row(ws, row, "氏名", str(record.get("氏名", "")))
    row += 1
    _write_info_row(ws, row, "作成日", str(record.get("作成日", "")))
    row += 1
    _write_info_row(ws, row, "支給日", str(record.get("支給日", "")))

    row += 2
    _write_block_header(ws, row, "【支給】")
    row += 1
    pay_items = _pay_items(record)
    for label, amount in pay_items:
        _write_row(ws, row, label, int(amount))
        row += 1
    _write_row(ws, row, "支給合計 A", int(record.get("A_給与総額", 0)), bold=True)
    row += 2

    _write_block_header(ws, row, "【控除】")
    row += 1
    deduct_items = [
        ("健康保険", _get_amount(record, "健康保険")),
        ("厚生年金", _get_amount(record, "厚生年金")),
        ("雇用保険", _get_amount(record, "雇用保険")),
        ("子ども・子育て支援金", _get_amount(record, "子ども・子育て支援金")),
        ("その他社会保険", _get_amount(record, "その他社会保険")),
        ("源泉所得税", _get_amount(record, "源泉所得税", "所得税")),
        ("市町村民税", _get_amount(record, "市町村民税")),
        ("その他控除", _get_amount(record, "その他控除")),
    ]
    for label, amount in deduct_items:
        _write_row(ws, row, label, int(amount))
        row += 1
    _write_row(ws, row, "控除合計", _deduction_total(record), bold=True)
    row += 2

    _write_block_header(ws, row, "【差引・計算内訳】")
    row += 1
    taxable_c = _get_amount(record, "C_課税対象給与")
    if taxable_c == 0:
        taxable_c = _get_amount(record, "A_給与総額")
    summary_items = [
        ("課税対象給与 C", taxable_c),
        ("社会保険料計 D", _get_amount(record, "D_社会保険料計")),
        ("差引控除後給与額 E", _get_amount(record, "E_差引控除後給与額")),
        ("控除計 F", _get_amount(record, "F_控除計")),
    ]
    for label, amount in summary_items:
        _write_row(ws, row, label, int(amount))
        row += 1
    _write_row(ws, row, "差引支給額 G", int(record.get("G_差引支給額", 0)), net=True)
    row += 2

    _write_block_header(ws, row, "【備考】")
    row += 1
    # 備考入力欄（A{row}:B{row+2}）を3行確保し、外周だけ罫線を付ける
    remark_start_row = row
    remark_end_row = row + 2
    ws.merge_cells(
        start_row=remark_start_row,
        start_column=1,
        end_row=remark_end_row,
        end_column=2,
    )
    remark = ws.cell(row=remark_start_row, column=1, value=str(record.get("備考", "") or ""))
    remark.font = NORMAL
    remark.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    _apply_outer_border(ws, remark_start_row, remark_end_row, 1, 2)
    ws.row_dimensions[remark_start_row].height = 18
    ws.row_dimensions[remark_start_row + 1].height = 18
    ws.row_dimensions[remark_start_row + 2].height = 18
    last_row = remark_end_row
    ws.print_area = f"A1:B{last_row}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True


def generate_payslip_workbook(records):
    wb = Workbook()
    wb.remove(wb.active)

    sheet_names = _unique_sheet_names(records)
    for record, sheet_name in zip(records, sheet_names):
        ws = wb.create_sheet(title=sheet_name)
        _build_sheet(ws, record)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def sample_record() -> dict:
    """印刷プレビュー確認用のサンプル1件。"""
    return {
        "作成日": "2026-04-25",
        "支給日": "2026-04-28",
        "月分": "4月分",
        "氏名": "山田太郎",
        "備考": "",
        "基本給": 250000,
        "時間外手当": 15000,
        "家族手当": 10000,
        "通勤手当": 5000,
        "資格手当": 0,
        "通信手当": 0,
        "A_給与総額": 280000,
        "C_課税対象給与": 280000,
        "健康保険": 12000,
        "厚生年金": 22000,
        "雇用保険": 1400,
        "子ども・子育て支援金": 800,
        "その他社会保険": 0,
        "D_社会保険料計": 36200,
        "E_差引控除後給与額": 243800,
        "源泉所得税": 5000,
        "市町村民税": 8000,
        "その他控除": 0,
        "F_控除計": 13000,
        "G_差引支給額": 230800,
    }
